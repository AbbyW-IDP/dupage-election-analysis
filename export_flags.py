"""
export_flags.py
---------------
Export unresolved contest name flags to an Excel workbook for review.

The workbook has two tabs:
  - "flags"          Unresolved flags with a Status column for your review
  - "known_contests" All normalized contest names currently in the registry

Workflow:
  1. python export_flags.py                  # produces flags_review.xlsx
  2. Open the workbook and set Status to "accepted" or fill in
     "Override Target" for flags you want to map to an existing contest
  3. python import_flags.py                  # reads the workbook and updates the DB

Status values:
  unreviewed   Default. Flag will be skipped on import.
  accepted     Accept the normalized name as a new contest.
  mapped       Map to an existing contest. Fill in "Override Target" with
               the exact normalized name from the known_contests tab.
  ignored      Flag is acknowledged but should not be loaded into the DB.

Usage:
    python export_flags.py [output.xlsx]
"""

import sys
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from dupage_elections.db import ElectionDatabase, DEFAULT_DB_PATH

DEFAULT_OUTPUT = Path("flags_review.xlsx")


def export_flags(
    db_path: Path = DEFAULT_DB_PATH,
    output_path: Path = DEFAULT_OUTPUT,
) -> None:
    with ElectionDatabase(db_path) as db:
        flags = db.get_unresolved_flags()
        known = sorted(db.get_known_contest_names())

    if not flags:
        print("No unresolved flags to export.")
        return

    # Build flags DataFrame
    flags_df = pd.DataFrame(flags)[["id", "year", "contest_name_raw", "contest_name"]]
    flags_df = flags_df.rename(columns={
        "id":               "Flag ID",
        "year":             "Year",
        "contest_name_raw": "Raw Name",
        "contest_name":     "Normalized Suggestion",
    })
    flags_df["Status"] = "unreviewed"
    flags_df["Override Target"] = ""
    flags_df["Notes"] = ""

    # Build known contests DataFrame
    known_df = pd.DataFrame({"Normalized Contest Name": known})

    print(f"Exporting {len(flags_df)} flags and {len(known_df)} known contest names...")

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        flags_df.to_excel(writer, sheet_name="flags", index=False)
        known_df.to_excel(writer, sheet_name="known_contests", index=False)

        _format_flags_sheet(writer.sheets["flags"], flags_df)
        _format_known_sheet(writer.sheets["known_contests"])

    print(f"Written to {output_path}")
    print()
    print("Next steps:")
    print("  1. Open the workbook and review the 'flags' tab")
    print("  2. Set Status to: accepted, mapped, or ignored")
    print("     For 'mapped', fill in 'Override Target' with a name from the 'known_contests' tab")
    print("  3. Run: python import_flags.py")


def _format_flags_sheet(ws, df: pd.DataFrame) -> None:
    """Apply formatting to the flags sheet."""
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Column widths
    widths = {
        "Flag ID": 10,
        "Year": 8,
        "Raw Name": 55,
        "Normalized Suggestion": 55,
        "Status": 14,
        "Override Target": 55,
        "Notes": 35,
    }
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(col_name, 20)

    # Freeze header row
    ws.freeze_panes = "A2"


def _format_known_sheet(ws) -> None:
    """Apply formatting to the known contests sheet."""
    header_fill = PatternFill("solid", fgColor="375623")
    header_font = Font(bold=True, color="FFFFFF")
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=1).font = header_font
    ws.column_dimensions["A"].width = 65
    ws.freeze_panes = "A2"


if __name__ == "__main__":
    output_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_OUTPUT
    db_path = Path(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_DB_PATH
    export_flags(db_path=db_path, output_path=output_path)
