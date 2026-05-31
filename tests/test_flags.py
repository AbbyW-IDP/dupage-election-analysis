"""
tests/test_flags.py
-------------------
Tests for flags.py: import_flags() and the remap_by_raw_name integration.

The interactive review_flags() function is not tested here -- it requires
stdin simulation and is covered by manual testing.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from src.election_analysis_generator.db import ElectionDatabase
from src.election_analysis_generator.flags import import_flags


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _seed_flagged(
    db: ElectionDatabase,
    raw: str,
    normalized: str,
    year: int = 2014,
    election_name: str = "2014 General Primary",
) -> int:
    """Insert a contest_results row with an unresolved flag, simulating a load
    that produced an unrecognized contest name.

    Returns the flag_id.
    """
    db._conn.execute(
        "INSERT OR IGNORE INTO contests (contest_name, is_legislation)"
        " VALUES (?, 0)",
        (normalized,),
    )
    db._conn.execute(
        "INSERT OR IGNORE INTO elections"
        " (name, year, summary_file, category, election_type)"
        " VALUES (?, ?, ?, ?, ?)",
        (election_name, year, f"{election_name}.csv", "General Primary", "midterm"),
    )
    eid = db._conn.execute(
        "SELECT id FROM elections WHERE name = ?", (election_name,)
    ).fetchone()[0]
    cid = db._conn.execute(
        "SELECT id FROM contests WHERE contest_name = ?", (normalized,)
    ).fetchone()[0]
    db._conn.execute(
        "INSERT INTO contest_results"
        " (contest_id, election_id, line_number, contest_name_raw,"
        "  contest_name, election_name, year, choice_name, party, total_votes)"
        " VALUES (?, ?, 1, ?, ?, ?, ?, ?, ?, ?)",
        (cid, eid, raw, normalized, election_name, year, "X", "DEM", 1000),
    )
    db._conn.execute(
        "INSERT INTO contest_flags (year, contest_name_raw, contest_name)"
        " VALUES (?, ?, ?)",
        (year, raw, normalized),
    )
    flag_id: int = db._conn.execute(
        "SELECT id FROM contest_flags WHERE contest_name_raw = ?", (raw,)
    ).fetchone()[0]
    db._conn.commit()
    return flag_id


def _make_xlsx(
    tmp_path: Path,
    rows: list[dict],
    known_names: list[str] | None = None,
) -> Path:
    """Write a minimal flags_review.xlsx with the given flag rows."""
    path = tmp_path / "flags_review.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]
    ws.title = "flags"
    ws.append([
        "Flag ID", "Year", "Raw Name", "Normalized Suggestion",
        "Status", "Override Target", "Notes",
    ])
    for row in rows:
        ws.append([
            row["flag_id"],
            row.get("year", 2014),
            row["raw"],
            row["normalized_suggestion"],
            row["status"],
            row.get("override_target", ""),
            row.get("notes", ""),
        ])
    ws2 = wb.create_sheet("known_contests")
    ws2.append(["Normalized Contest Name"])
    for name in (known_names or ["FOR ATTORNEY GENERAL", "FOR COUNTY CLERK"]):
        ws2.append([name])
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture
def seeded_db(db: ElectionDatabase) -> ElectionDatabase:
    """DB with FOR ATTORNEY GENERAL pre-registered as a known canonical name."""
    db._conn.execute(
        "INSERT OR IGNORE INTO contest_name_registry (contest_name, first_seen_year)"
        " VALUES (?, 0)",
        ("FOR ATTORNEY GENERAL",),
    )
    db._conn.commit()
    return db


# ---------------------------------------------------------------------------
# accepted -- name unchanged
# ---------------------------------------------------------------------------


class TestImportFlagsAccepted:

    def test_accepted_unchanged_resolves_flag(self, seeded_db, tmp_path):
        """accepted with unchanged name resolves the flag."""
        # Arrange
        flag_id = _seed_flagged(
            seeded_db, "FOR ATTORNEY GENERAL", "FOR ATTORNEY GENERAL"
        )
        xlsx = _make_xlsx(tmp_path, [{
            "flag_id": flag_id, "raw": "FOR ATTORNEY GENERAL",
            "normalized_suggestion": "FOR ATTORNEY GENERAL", "status": "accepted",
        }])

        # Act
        counts = import_flags(seeded_db, xlsx)

        # Assert
        assert counts["accepted"] == 1
        assert seeded_db.get_unresolved_flags() == []

    def test_accepted_unchanged_leaves_contest_results_intact(
        self, seeded_db, tmp_path
    ):
        """accepted with unchanged name does not alter contest_results rows."""
        # Arrange
        flag_id = _seed_flagged(
            seeded_db, "FOR ATTORNEY GENERAL", "FOR ATTORNEY GENERAL"
        )
        xlsx = _make_xlsx(tmp_path, [{
            "flag_id": flag_id, "raw": "FOR ATTORNEY GENERAL",
            "normalized_suggestion": "FOR ATTORNEY GENERAL", "status": "accepted",
        }])

        # Act
        import_flags(seeded_db, xlsx)

        # Assert
        names = [
            r[0]
            for r in seeded_db._conn.execute(
                "SELECT DISTINCT contest_name FROM contest_results"
            ).fetchall()
        ]
        assert names == ["FOR ATTORNEY GENERAL"]

    # -----------------------------------------------------------------
    # accepted -- name changed (cross-year rename)
    # -----------------------------------------------------------------

    def test_accepted_changed_name_updates_contest_results(
        self, seeded_db, tmp_path
    ):
        """accepted with a corrected Normalized Suggestion remaps contest_results."""
        # Arrange
        flag_id = _seed_flagged(
            seeded_db,
            "Attorney General, State of Illinois - D*",
            "ATTORNEY GENERAL, STATE OF ILLINOIS",
        )
        xlsx = _make_xlsx(tmp_path, [{
            "flag_id": flag_id,
            "raw": "Attorney General, State of Illinois - D*",
            "normalized_suggestion": "FOR ATTORNEY GENERAL",
            "status": "accepted",
        }])

        # Act
        import_flags(seeded_db, xlsx)

        # Assert
        names = [
            r[0]
            for r in seeded_db._conn.execute(
                "SELECT DISTINCT contest_name FROM contest_results"
            ).fetchall()
        ]
        assert names == ["FOR ATTORNEY GENERAL"]

    def test_accepted_changed_name_resolves_flag(self, seeded_db, tmp_path):
        """accepted with a corrected name resolves the flag."""
        # Arrange
        flag_id = _seed_flagged(
            seeded_db,
            "Attorney General, State of Illinois - D*",
            "ATTORNEY GENERAL, STATE OF ILLINOIS",
        )
        xlsx = _make_xlsx(tmp_path, [{
            "flag_id": flag_id,
            "raw": "Attorney General, State of Illinois - D*",
            "normalized_suggestion": "FOR ATTORNEY GENERAL",
            "status": "accepted",
        }])

        # Act
        import_flags(seeded_db, xlsx)

        # Assert
        assert seeded_db.get_unresolved_flags() == []

    def test_accepted_resolves_all_years_for_raw_name(self, seeded_db, tmp_path):
        """accepted resolves all flags for the raw name, not just the workbook row."""
        # Arrange: same raw name in two different election years -> two flags
        raw = "Attorney General, State of Illinois - D*"
        flag_id_2014 = _seed_flagged(
            seeded_db, raw, "ATTORNEY GENERAL, STATE OF ILLINOIS",
            year=2014, election_name="2014 General Primary",
        )
        _seed_flagged(
            seeded_db, raw, "ATTORNEY GENERAL, STATE OF ILLINOIS",
            year=2016, election_name="2016 General Primary",
        )
        # Workbook only includes the 2014 row
        xlsx = _make_xlsx(tmp_path, [{
            "flag_id": flag_id_2014,
            "raw": raw,
            "normalized_suggestion": "FOR ATTORNEY GENERAL",
            "status": "accepted",
        }])

        # Act
        import_flags(seeded_db, xlsx)

        # Assert: both years resolved, not just the one in the workbook
        assert seeded_db.get_unresolved_flags() == []

    def test_accepted_changed_name_stores_override(self, seeded_db, tmp_path):
        """accepted with a corrected name stores an override for future loads."""
        # Arrange
        raw = "Attorney General, State of Illinois - D*"
        flag_id = _seed_flagged(
            seeded_db, raw, "ATTORNEY GENERAL, STATE OF ILLINOIS"
        )
        xlsx = _make_xlsx(tmp_path, [{
            "flag_id": flag_id, "raw": raw,
            "normalized_suggestion": "FOR ATTORNEY GENERAL",
            "status": "accepted",
        }])

        # Act
        import_flags(seeded_db, xlsx)

        # Assert
        assert seeded_db.get_overrides().get(raw) == "FOR ATTORNEY GENERAL"


# TODO: change the source code so the test is relevant or remove the test.
    # def test_accepted_changed_name_deletes_old_contests_row(
    #     self, seeded_db, tmp_path
    # ):
    #     """The orphaned contests row for the old name is cleaned up."""
    #     # Arrange
    #     flag_id = _seed_flagged(
    #         seeded_db,
    #         "Attorney General, State of Illinois - D*",
    #         "ATTORNEY GENERAL, STATE OF ILLINOIS",
    #     )
    #     xlsx = _make_xlsx(tmp_path, [{
    #         "flag_id": flag_id,
    #         "raw": "Attorney General, State of Illinois - D*",
    #         "normalized_suggestion": "FOR ATTORNEY GENERAL",
    #         "status": "accepted",
    #     }])

    #     # Act
    #     import_flags(seeded_db, xlsx)

    #     # Assert
    #     old_row = seeded_db._conn.execute(
    #         "SELECT id FROM contests WHERE contest_name = ?",
    #         ("ATTORNEY GENERAL, STATE OF ILLINOIS",),
    #     ).fetchone()
    #     assert old_row is None

    def test_accepted_registers_canonical_name_in_registry(
        self, db, tmp_path
    ):
        """accepted registers the (possibly corrected) name in contest_name_registry."""
        # Arrange
        flag_id = _seed_flagged(db, "NEW CONTEST", "NEW CONTEST")
        xlsx = _make_xlsx(tmp_path, [{
            "flag_id": flag_id, "raw": "NEW CONTEST",
            "normalized_suggestion": "NEW CONTEST", "status": "accepted",
        }], known_names=["NEW CONTEST"])

        # Act
        import_flags(db, xlsx)

        # Assert
        assert "NEW CONTEST" in db.get_known_contest_names()


# ---------------------------------------------------------------------------
# mapped
# ---------------------------------------------------------------------------


class TestImportFlagsMapped:

    def test_mapped_updates_contest_results(self, seeded_db, tmp_path):
        """mapped remaps contest_results rows to the Override Target."""
        # Arrange
        flag_id = _seed_flagged(
            seeded_db,
            "Attorney General, State of Illinois - D*",
            "ATTORNEY GENERAL, STATE OF ILLINOIS",
        )
        xlsx = _make_xlsx(tmp_path, [{
            "flag_id": flag_id,
            "raw": "Attorney General, State of Illinois - D*",
            "normalized_suggestion": "ATTORNEY GENERAL, STATE OF ILLINOIS",
            "status": "mapped",
            "override_target": "FOR ATTORNEY GENERAL",
        }])

        # Act
        counts = import_flags(seeded_db, xlsx)

        # Assert
        assert counts["mapped"] == 1
        names = [
            r[0]
            for r in seeded_db._conn.execute(
                "SELECT DISTINCT contest_name FROM contest_results"
            ).fetchall()
        ]
        assert names == ["FOR ATTORNEY GENERAL"]

    def test_mapped_resolves_flag(self, seeded_db, tmp_path):
        """mapped resolves the flag."""
        # Arrange
        flag_id = _seed_flagged(
            seeded_db,
            "Attorney General, State of Illinois - D*",
            "ATTORNEY GENERAL, STATE OF ILLINOIS",
        )
        xlsx = _make_xlsx(tmp_path, [{
            "flag_id": flag_id,
            "raw": "Attorney General, State of Illinois - D*",
            "normalized_suggestion": "ATTORNEY GENERAL, STATE OF ILLINOIS",
            "status": "mapped",
            "override_target": "FOR ATTORNEY GENERAL",
        }])

        # Act
        import_flags(seeded_db, xlsx)

        # Assert
        assert seeded_db.get_unresolved_flags() == []

    def test_mapped_stores_override(self, seeded_db, tmp_path):
        """mapped stores an override so future loads use the canonical name."""
        # Arrange
        raw = "Attorney General, State of Illinois - D*"
        flag_id = _seed_flagged(
            seeded_db, raw, "ATTORNEY GENERAL, STATE OF ILLINOIS"
        )
        xlsx = _make_xlsx(tmp_path, [{
            "flag_id": flag_id, "raw": raw,
            "normalized_suggestion": "ATTORNEY GENERAL, STATE OF ILLINOIS",
            "status": "mapped",
            "override_target": "FOR ATTORNEY GENERAL",
        }])

        # Act
        import_flags(seeded_db, xlsx)

        # Assert
        assert seeded_db.get_overrides().get(raw) == "FOR ATTORNEY GENERAL"

    def test_mapped_blank_override_target_counts_as_error(
        self, seeded_db, tmp_path
    ):
        """mapped with blank Override Target counts as an error."""
        # Arrange
        flag_id = _seed_flagged(
            seeded_db, "OLD NAME", "OLD NAME"
        )
        xlsx = _make_xlsx(tmp_path, [{
            "flag_id": flag_id, "raw": "OLD NAME",
            "normalized_suggestion": "OLD NAME",
            "status": "mapped",
            "override_target": "",
        }])

        # Act
        counts = import_flags(seeded_db, xlsx)

        # Assert
        assert counts["errors"] == 1
        assert counts["mapped"] == 0

    def test_mapped_unknown_override_target_counts_as_error(
        self, seeded_db, tmp_path
    ):
        """mapped with Override Target not in known_contests counts as an error."""
        # Arrange
        flag_id = _seed_flagged(seeded_db, "OLD NAME", "OLD NAME")
        xlsx = _make_xlsx(tmp_path, [{
            "flag_id": flag_id, "raw": "OLD NAME",
            "normalized_suggestion": "OLD NAME",
            "status": "mapped",
            "override_target": "DOES NOT EXIST",
        }])

        # Act
        counts = import_flags(seeded_db, xlsx)

        # Assert
        assert counts["errors"] == 1


# ---------------------------------------------------------------------------
# ignored / skipped
# ---------------------------------------------------------------------------


class TestImportFlagsIgnoredAndSkipped:

    def test_ignored_resolves_flag_without_registering(self, db, tmp_path):
        """ignored resolves the flag but does not register the name."""
        # Arrange
        flag_id = _seed_flagged(db, "BALLOT MEASURE A", "BALLOT MEASURE A")
        xlsx = _make_xlsx(tmp_path, [{
            "flag_id": flag_id, "raw": "BALLOT MEASURE A",
            "normalized_suggestion": "BALLOT MEASURE A", "status": "ignored",
        }], known_names=[])

        # Act
        counts = import_flags(db, xlsx)

        # Assert
        assert counts["ignored"] == 1
        assert db.get_unresolved_flags() == []
        assert "BALLOT MEASURE A" not in db.get_known_contest_names()

    def test_unreviewed_rows_are_skipped(self, db, tmp_path):
        """unreviewed rows are counted as skipped and left unresolved."""
        # Arrange
        flag_id = _seed_flagged(db, "SOME CONTEST", "SOME CONTEST")
        xlsx = _make_xlsx(tmp_path, [{
            "flag_id": flag_id, "raw": "SOME CONTEST",
            "normalized_suggestion": "SOME CONTEST", "status": "unreviewed",
        }], known_names=[])

        # Act
        counts = import_flags(db, xlsx)

        # Assert
        assert counts["skipped"] == 1
        assert len(db.get_unresolved_flags()) == 1


# ---------------------------------------------------------------------------
# File-level error handling
# ---------------------------------------------------------------------------


class TestImportFlagsErrors:

    def test_raises_file_not_found(self, db, tmp_path):
        """Raises FileNotFoundError when the workbook does not exist."""
        # Arrange
        missing = tmp_path / "nonexistent.xlsx"

        # Act / Assert
        with pytest.raises(FileNotFoundError):
            import_flags(db, missing)

    def test_raises_value_error_for_missing_columns(self, db, tmp_path):
        """Raises ValueError when required columns are absent from the workbook."""
        # Arrange
        path = tmp_path / "bad.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.worksheets[0]
        ws.title = "flags"
        ws.append(["Wrong Column"])
        wb.save(path)

        # Act / Assert
        with pytest.raises(ValueError, match="Missing columns"):
            import_flags(db, path)


# ---------------------------------------------------------------------------
# Schema migration
# ---------------------------------------------------------------------------


def test_contest_flags_has_source_columns(tmp_path):
    db_path = tmp_path / "test.db"
    with ElectionDatabase(db_path) as db:
        cols = {
            row[1]
            for row in db._conn.execute(
                "PRAGMA table_info(contest_flags)"
            ).fetchall()
        }
    assert "source_file" in cols
    assert "source_tab" in cols
    assert "source_row" in cols


def test_write_flags_stores_source_info(tmp_path):
    import pandas as pd

    db_path = tmp_path / "test.db"
    with ElectionDatabase(db_path) as db:
        df = pd.DataFrame(
            {
                "contest_name_raw": ["RAW CONTEST A", "RAW CONTEST A"],
                "contest_name": ["raw contest a", "raw contest a"],
            }
        )
        # index 0 and 1 -> first occurrence of "RAW CONTEST A" is index 0 -> row = 0 + 2 = 2
        db._write_flags(
            df,
            year=2024,
            known=set(),
            source_file="results.csv",
            source_tab=None,
        )
        row = db._conn.execute(
            "SELECT source_file, source_tab, source_row FROM contest_flags"
        ).fetchone()
    assert row["source_file"] == "results.csv"
    assert row["source_tab"] is None
    assert row["source_row"] == 2


def test_csv_load_stores_source_file_in_flags(tmp_path):
    import pandas as pd
    from election_analysis_generator.models import Election
    from datetime import date

    db_path = tmp_path / "test.db"

    df = pd.DataFrame({
        "contest_name_raw": ["BRAND NEW CONTEST"],
        "choice_name": ["Candidate A"],
        "party": ["DEM"],
        "total_votes": [100],
    })

    election = Election(
        id=None,
        name="2024 Test",
        year=2024,
        election_date=date(2024, 11, 5),
        category="general",
        election_type="presidential",
        summary_file="results.csv",
    )

    with ElectionDatabase(db_path) as db:
        db.insert_election_with_file(election, df, "results.csv")
        flags = db.get_unresolved_flags()

    assert len(flags) == 1
    assert flags[0]["source_file"] == "results.csv"
    assert flags[0]["source_tab"] is None
    assert flags[0]["source_row"] == 2


def test_excel_load_stores_source_tab_in_flags(tmp_path):
    import openpyxl
    from src.election_analysis_generator.loader import LoadSummary, LoadPrecinctDetail

    db_path = tmp_path / "test.db"

    _CONTEST_UNRECOGNIZED = "SOME TOTALLY UNKNOWN CONTEST (Vote For 1)"
    _CANDIDATE = "Jane Smith"
    _CONTEST_KNOWN = "FOR GOVERNOR (Vote For 1)"

    CSV_HEADER = (
        "line number,contest name,choice name,party name,"
        "total votes,percent of votes,registered voters,ballots cast,"
        "num Precinct total,num Precinct rptg,over votes,under votes"
    )

    with ElectionDatabase(db_path) as db:
        # Seed a summary election so we have a valid election.id
        csv_path = tmp_path / "2026-general-primary.csv"
        csv_path.write_text(
            CSV_HEADER + "\n"
            f"1,{_CONTEST_KNOWN},{_CANDIDATE},D,5000,100.0,50000,10000,10,10,0,0\n"
        )
        config = {
            "name": "2026 General Primary",
            "year": 2026,
            "summary_file": csv_path.name,
            "election_date": "2026-04-07",
        }
        election, _ = LoadSummary(db).load_csv(csv_path, config)

        # Build a .xlsx with one sheet named "DEM GOVERNOR" containing
        # a contest name that is NOT in the registry -- this will produce a flag.
        xlsx_path = tmp_path / "detail.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "DEM GOVERNOR"
        ws.append([_CONTEST_UNRECOGNIZED])
        ws.append([None, None, _CANDIDATE])
        ws.append([])
        ws.append(["Addison 001", 1200, 50, 100, 300, 10, 460])
        wb.save(xlsx_path)

        LoadPrecinctDetail(db).load_detail_excel(xlsx_path, election)

        flags = db.get_unresolved_flags()

    xlsx_flags = [f for f in flags if f["source_file"] == xlsx_path.name]
    assert len(xlsx_flags) == 1
    assert xlsx_flags[0]["source_tab"] == "DEM GOVERNOR"


def test_resolve_flag_by_raw_name(tmp_path):
    db_path = tmp_path / "test.db"
    with ElectionDatabase(db_path) as db:
        db._conn.execute(
            "INSERT INTO contest_flags (year, contest_name_raw, contest_name) VALUES (?,?,?)",
            (2024, "RAW A", "raw a"),
        )
        db._conn.commit()
        db.resolve_flag_by_raw_name("RAW A")
        db._conn.commit()
        remaining = db.get_unresolved_flags()
    assert remaining == []
